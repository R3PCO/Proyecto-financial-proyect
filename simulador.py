 # -*- coding: utf-8 -*-
"""
Created on Fri Oct 30 08:47:21 2020

@author: PERSONAL
"""

from tkinter import *
from tkinter import filedialog as FileDialog
from tkinter import messagebox as MessageBox
from tkinter import ttk
import openpyxl

#ventana principal
root = Tk()
root.title("Mi simulador")


ruta = "" #almacenar la ruta donde se ubicará el ficherot
global wb
global texto

wb = openpyxl.Workbook()

def make_hojas():
        
        frame= Frame(texto)
        frame.pack()
        frame.config(bg="lightblue")     
        frame.config(width=200,height=150) 
        frame.pack(fill="both", expand = "1")
        
       #def test():
            #MessageBox.showinfo(nombre) # título, mensaje
      
        nh=int(numero_hojas.get())    
        
        def nombrar_hoja():
            
            global hoja
            nombre = nombre_hoja.get()    
            print(nombre)
            hoja= wb.create_sheet(nombre)
            print(hoja)
            return hoja
            
        
        c=0 
        
        while c <= nh-1:
             """
             if c==0:
                  nombre_hdefect =StringVar()
                  label = Label(frame, text = "cambia el nombre de la hoja creada por defecto").pack()
                  Entry(frame, texvariable = nombre_hdefect).pack()
                  nombrehdefect = nombre_hdefect.get()
                  hojadefec = wb.active
                  hojadefec.title = nombrehdefect   
             else:
             """
             nombre_hoja= StringVar()
             label=Label(frame, text= "incluye el nombre de la hoja").pack()
             Entry(frame, textvariable= nombre_hoja).pack()
             Button(frame, text="nombrar_hoja", command= nombrar_hoja).pack()     
             c+=1         
        else:        
              print("final iteracion")
              print(wb.sheetnames)
              
             
              
def crear_excel():
  
    global ruta
    global numero_hojas
  
    mensaje.set('Nuevo archivo')
    texto.delete(1.0, END)
    
    
    frame= Frame (texto)
    frame.pack()
    frame.config(bg="lightblue")     
    frame.config(width=200,height=150) 
    frame.pack(fill="both", expand = "1")
    numero_hojas=StringVar()

    
    label =Label(frame, text ="incluye el numero de hojas").pack()
    Entry(frame, justify=CENTER, textvariable = numero_hojas).pack()
    Button(frame, text="hacer hojas", command=make_hojas).pack()

 
    return

"""     


    se accede a la primera hoja
    a1 = hoja["A1"] # se accede a la columan a fila 1
    print(a1.value)
    # 1.- Asignando el valor directamente a la celda
    hoja["A1"] = 10
    a1 = hoja["A1"]
    print(a1.value)
    a1.value =50
    print(a1.value)
   
    
    #incluir una hoja con los parametros
    
    parametros =[ # Habitaciones
                ("Tarifa Promedio",120), 
                ("Unidad Hotelera", 310.000), 
                ("Porcentaje de Ocupación",1.5),
                ("Enero",100)]

    hoja.append(("parametro","valor"))
    
    for parametro in parametros:
        hoja.append(parametro)
    
    wb.save("simulador.xlsx")

"""
    
def abrir(): 
    mensaje.set('abrir archivo')
    ruta = FileDialog.askopenfilename(
        initialdir='e:/Users/PERSONAL/Documents/Curso Python mundo real ejercicios/001- leer archivos excel/input',
        filetypes=(  # Es una tupla con un elemento
            ("Ficheros de excel", "*.xlsx"),  
        ), 
        title="Abrir un archivo archivo excel."
    )
    if ruta != "":  
            fichero = open(ruta, 'r')
            nombre_completo = os.path.join(path,nombre_archivo)
  


def guardar():
    mensaje.set('Guardar fichero')
    pass
def guardar_como():
    print("Guardar fichero como")
    pass


def inclusion():
    pass


def calculos():
    pass

def resultados():
    pass

#Menú superior
menubar = Menu(root)
root.config(menu = menubar)


filemenu = Menu(menubar, tearoff=0)
menubar.add_cascade(label="Archivo", menu=filemenu)

filemenu.add_command(label="Abrir", command = abrir)
filemenu.add_command(label="Guardar", command = guardar)
filemenu.add_command(label="Guardar_como", command = guardar_como)
filemenu.add_separator()
filemenu.add_command(label="Salir", command=root.quit)
filemenu.add_separator()
submenucrear=Menu(menubar,tearoff = 0)
filemenu.add_cascade(label = "crear_excel", menu = submenucrear)
submenucrear.add_command(label="crear", command = crear_excel)




submenucrear.add_command(label="hacer_hojas", command = make_hojas)
submenucrear.add_command(label="calulos", command = calculos)




filemenu2= Menu(menubar, tearoff=0)
menubar.add_cascade(label="aplicacion", menu=filemenu2)

filemenu2.add_command(label="introducir_datos", command = inclusion)
filemenu2.add_command(label="calulos", command = calculos)
filemenu2.add_separator()
filemenu2.add_command(label="mostrar_resultados", command = resultados)



texto = Text(root)
texto.pack(fill='both', expand=1)
texto.config(padx=6, pady=4, bd=0, font=("Consolas", 12))

root.config(menu = menubar)
# Menu y bucle de la aplicación
#root.config(menu = menubar2)


mensaje = StringVar()
mensaje.set('Bienvenido a tu simulador')
monitor = Label(root, textvar=mensaje, justify='right')
monitor.pack(side='left')

root.mainloop()